# Copyright (c) 2020, Sistem Koperasi
from __future__ import unicode_literals

import frappe

import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Protection
from openpyxl import load_workbook
from six import BytesIO, string_types
from openpyxl.utils import get_column_letter

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
# return xlsx file object
def make_xlsx(data, sheet_name, protect_area=None, wb=None ):

	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)

	row1 = ws.row_dimensions[1]
	row1.font = Font(name='Calibri',bold=True)


	for row in data:
		clean_row = []
		column_widths = []
		for i, cell in enumerate(row):

			if len(column_widths) > i:
				if len(cell) > column_widths[i]:
					column_widths[i] = len(cell)
			else:
				if isinstance(cell, string_types):
					cell_length = len(cell)
					column_widths += [cell_length+1 if cell_length > 11 else 11]
				else:
					column_widths += [10]

		for item in row:
			if isinstance(item, string_types) and (sheet_name not in ['Data Import Template', 'Data Export']):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, string_types) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)

			clean_row.append(value)

		ws.append(clean_row)

	for i, column_width in enumerate(column_widths):
		ws.column_dimensions[get_column_letter(i+1)].width = column_width

	if isinstance(protect_area, list):
		# Set Protection on cell not intended to be edited
		for col in ws.iter_cols(min_row=protect_area[0], min_col = protect_area[1], max_col=protect_area[2], max_row=len(data)+1):
			for cell in col:
				cell.number_format  = "##"
				cell.protection = Protection(locked=False)

		ws.protection.sheet = True
		ws.protection.password = 'Suk4suk4'
		ws.protection.enable()
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


def handle_html(data):
	# return if no html tags found
	data = frappe.as_unicode(data)

	if '<' not in data:
		return data
	if '>' not in data:
		return data

	from html2text import HTML2Text

	h = HTML2Text()
	h.unicode_snob = True
	h = h.unescape(data or "")

	obj = HTML2Text()
	obj.ignore_links = True
	obj.body_width = 0

	try:
		value = obj.handle(h)
	except Exception:
		# unable to parse html, send it raw
		return data

	value = ", ".join(value.split('  \n'))
	value = " ".join(value.split('\n'))
	value = ", ".join(value.split('# '))

	return value

def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None, sheet=None):
	if file_url:
		_file = frappe.get_doc("File", {"file_url": file_url})
		filename = _file.get_full_path()
	elif fcontent:
		from io import BytesIO
		filename = BytesIO(fcontent)
	elif filepath:
		filename = filepath
	else:
		return

	rows = []
	wb1 = load_workbook(filename=filename, read_only=True, data_only=True)
	if sheet:
		ws1 = wb1[sheet]
	else:
		ws1 = wb1.active
	for row in ws1.iter_rows():
		tmp_list = []
		for cell in row:
			tmp_list.append(cell.value)
		rows.append(tmp_list)
	return rows

def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	rows = []
	for i in range(sheet.nrows):
		rows.append(sheet.row_values(i))
	return rows

def build_xlsx_response(data, filename, protect_area):
	xlsx_file = make_xlsx(data, filename, protect_area)
	# write out response as a xlsx type
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'
